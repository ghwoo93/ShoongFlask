
from flask import Flask,render_template,request,jsonify,session
from flask_restful import Api
from flask_cors import CORS
from places.matzip import Matzip
from places.janggwan import JangGwan
from jatooroute.jatoo import Jatoo

import os
import dialogflow
import uuid#세션아이디로 사용
from settings.config import DIALOG_CONFIG#프로젝트 아이디/API키가 설정된 모듈 import

#플라스크 앱 생성
# Flask 객체 선언, 파라미터로 어플리케이션 패키지의 이름을 넣어줌.
app = Flask(__name__)
#JSON 응답 한글 처리
app.config['JSON_AS_ASCII']=False
#CORS에러 처리
# 데코레이터 이용, '/hello' 경로에 클래스 등록
CORS(app)
#플라스크 앱을 인자로 하여 Api객체 생성:클래스와 URI매핑
# Flask 객체에 Api 객체 등록
api = Api(app)

#별 추가 
app.config['JSON_AS_ASCII'] = False
#session을 위한 시크릿 키 설정:임의의 문자열- 세션에 값 설정시 반드시 필요
app.secret_key='dfgsdfg@$shfdsg&sfdsafh'

#GOOGLE_APPLICATION_CREDENTIALS키워드에 대한 설명:https://dialogflow.com/docs/reference/v2-auth-setup
#                                                ->https://cloud.google.com/docs/authentication/getting-started(2021.01.13기준)
#환경 변수 GOOGLE_APPLICATION_CREDENTIALS를 설정하여 애플리케이션 코드에 사용자 인증 정보를 제공
#API키를 환경변수로 등록
os.environ['GOOGLE_APPLICATION_CREDENTIALS']=DIALOG_CONFIG['GOOGLE_APPLICATION_CREDENTIALS']
#별 추가  끝 


#요청을 처리할 클래스와 요청 uri 매핑(라우팅)
#Api객체.add_resource(클래스명,'/요청url')
#/todos/<todo_id> url 패턴이면
#get방식이면 todo_id로 조회
#delete방식이면 todo_id로 삭제
#put방식이면 todo_id로 수정
#api.add_resource(Matzip,'/shoong/<lat>/<lng>')
api.add_resource(Matzip,'/places/matzip')
api.add_resource(JangGwan,'/places/janggwan')
api.add_resource(Jatoo,'/jatoo')
#/todos 로 요청시 get방식이면 전체조회
#                post방식이면 할일 등록
#api.add_resource(TodoList,'/todos')
#api.add_resource(Upload,'/upload')

if __name__ =='__main__':
    app.run(port=10004,host='0.0.0.0')



#별추가 : 테스트완료 추후 완료후 추가 하겠음
@app.route('/message',methods=['GET'])
def handleMessage():#사용자 UI(Client App)에서 보낸 대화를 받는 함수
              #받은 대화는 다시 DialogFlow로 보낸다

    session['session_id'] = str(uuid.uuid4())#다른 어플리케이션의 UI사용시
    message= request.values.get('message')


    print('사용자 UI(Client App)에서 입력한 메시지:',message)
    #프로젝트 아이디 가져오기
    project_id = DIALOG_CONFIG.get('PROJECT_ID')
    #플라스크앱이  다얼로그 플로우로부터 받은 응답
    fulfillmentText = response_from_dialogflow(project_id,session['session_id'],message,'ko')

    #다이얼로그로부터 받은 응답을 클라이언트 App(사용자 UI)에 전송
    return jsonify({'message':fulfillmentText})


def response_from_dialogflow(project_id, session_id, message, language_code):
    # step1. DialogFlow와 사용자가 상호작용할 세션 클라이언트 생성
    session_client = dialogflow.SessionsClient()
    session_path = session_client.session_path(project_id, session_id)
    # projects/프로젝트아이디/agent/sessions/세션아이디 로 생성된다
    print('[session_path]', session_path, sep='\n')
    if message:  # 사용자가 대화를 입력한 경우.대화는 utf-8로 인코딩된 자연어.256자를 넘어서는 안된다
        # step2.사용자 메시지(일반 텍스트)로 TextInput생성
        text_input = dialogflow.types.TextInput(text=message, language_code=language_code)
        print('[text_input]', text_input, sep='\n')
        '''
        text : '사용자가 입력한 대화'
        language_code :'ko'        
        '''
        # step 3. 생성된 TextInput객체로 QueryInput객체 생성(DialogFlow로 전송할 질의 생성)
        query_input = dialogflow.types.QueryInput(text=text_input)
        print('[query_input]', query_input, sep='\n')
        '''
        text {
                text : '사용자가 입력한 대화'
                language_code :'ko'
             }        
        '''
        # step 4. DialogFlow로 SessionsClient객체.detect_intent()메소드로
        #        QueryInput객체를 보내고 다시 봇 응답(Responses섹션에 등록한 대화)을 받는다
        #        즉 A DetectIntentResponse instance반환
        '''
        PermissionDenied : 403 IAM 권한 에러시
        1. GOOGLE DEVELOPER CONSOLE로 검색후 
           좌측 상단 Google APIs 옆 프로젝트 목록에서 해당 프로젝트 선택 
        2. 좌측의 사용자 인증 정보 클릭
        3. 서비스 계정의 이메일 클릭->IAM 및 관리자
        4. IAM 및 관리자 페이지의 좌측의 IAM메뉴 클릭
        5. 상단의  추가 메뉴 클릭
        6. 구성원 항목에는 서비스 계정 생성시 다운받은 mytravelbot-vbem-cf60a98145b6.json파일의 
           "client_email"키의 값을 복사하여 넣는다
           역할 항목에는 "소유자"를 선택하여 넣는다.    

        '''
        response = session_client.detect_intent(session=session_path, query_input=query_input)
        print('[response]', response, sep='\n')
        print('[type(response)]', type(response), sep='\n')  # DetectIntentResponse타입

    return response.query_result.fulfillment_text  # 다이얼로그플로우 봇이 응답한 텍스트


# 아래 웹 후크용 메소드는 추가적으로 나만의 응답(데이타베이스에서 읽어오거나)
# 을 구성하고자 할때 사용.웹 후크를 사용하지 않아도 우리가 다이얼로그 플랫폼에
# 등록한 사용자 질의문과 응답으로도 충분히 나만의 챗봇을 만들 수 있다

# 웹 후크 서비스 : 즉 다이얼로그 플로우가 인텐트 매칭후
# 아래 API서비스(웹 후크)를 POST로 요청한다
# 전제조건
# 1. 웹 후크를 적용할 인텐트 선택후 fulfillment메뉴에서 enable설정
# 2. 해당 봇의 죄측 메뉴인 fullfillments탭에서 아래 url을  등록(loccalhost 및 http는 불가)

@app.route('/webhook_rpa', methods=['POST'])
def webhook():  # fulfillment를 enable로 설정한 인텐트로 진입했을때 DialogFlow가 이 URL 요청
    # 다이얼로그 플로우에서 json으로 응답을 보낸다
    webhook_response = request.get_json(force=True)
    print('[webhook_response]', webhook_response, sep='\n')
    # 아래는 챗봇 UI에 사용자가 입력한 Full text
    # query = webhook_response['queryResult']['queryText']#사용자 입력분 예]크롬 실행해 주세요
    # 아래는 엔터티 즉 파라미터명으로  값 추출
    # 대표 엔터티명으로 비교하면된다(그럼 모든 동의어도 처리가 된다)
    # 아래에서 'program'은 개발자 정의 엔터티
    program = webhook_response['queryResult']['parameters']['program']
    if '엑셀' in program:
        # https://openpyxl.readthedocs.io/en/stable/
        from openpyxl.workbook import Workbook
        wb = Workbook()
        sheet1 = wb['Sheet']
        sheet1.title = '오늘 할일'
        sheet1['A1'] = '오늘의 할 일 리스트'
        sheet1.append(['1.', '이메일 보내기'])
        sheet1.append(['2.', '고객명단 정리하기'])

        sheet2 = wb.create_sheet('내일 할일')
        sheet2.cell(row=1, column=1, value='내일의 할 일 리스트')
        # 작성 내용 엑셀 파일로 저장
        wb.save('todos.xlsx')
        # 엑셀 실행
        os.startfile('todos.xlsx')
        reply = {'fulfillmentText': '엑셀 실행합니다'}  # DialogFlow에 JSON으로 응답(키값은 반드시 'fulfillmentText')

    elif '브라우저' == program:
        import webbrowser
        webbrowser.open_new('https://www.google.com')
        reply = {'fulfillmentText': '브라우저 실행합니다'}
    else:
        reply = {'fulfillmentText': 'I cannot execute {}'.format(program)}

    return jsonify(reply)

#별 추가 테스트 끝